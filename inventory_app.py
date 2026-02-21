import os
import sqlite3
from datetime import datetime

from PySide6.QtGui import QFont
from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QTabWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QMessageBox, QFormLayout, QSpinBox, QTextEdit, QComboBox, QFileDialog
)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import barcode
from barcode.writer import ImageWriter
from PIL import Image, ImageDraw, ImageFont


APP_TITLE = "備品在庫管理"
BASE_DIR = r"\\dionas\管理部\備品管理\システム関連"
os.makedirs(BASE_DIR, exist_ok=True)
DB_FILE = os.path.join(BASE_DIR, "inventory.db")
LABEL_DIR = os.path.join(BASE_DIR, "labels")
os.makedirs(LABEL_DIR, exist_ok=True)


# ======== データベース ========
class DB:
    def __init__(self, path: str = DB_FILE):
        self.path = path
        self.conn = sqlite3.connect(self.path, timeout=30)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA busy_timeout = 30000;")
        try:
            self.conn.execute("PRAGMA journal_mode = WAL;")
        except Exception:
            self.conn.execute("PRAGMA journal_mode = DELETE;")
        self.conn.execute("PRAGMA synchronous = NORMAL;")
        self._init_schema()

    def _init_schema(self):
        cur = self.conn.cursor()
        cur.execute("""
        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            location TEXT,
            unit TEXT,
            safety_stock INTEGER DEFAULT 0,
            note TEXT,
            is_active INTEGER DEFAULT 1
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS stock (
            item_id INTEGER PRIMARY KEY,
            qty INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY(item_id) REFERENCES items(id)
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT NOT NULL,
            type TEXT NOT NULL,
            item_id INTEGER NOT NULL,
            qty INTEGER NOT NULL,
            supplier TEXT,          -- 入庫：購入先
            user TEXT,              -- 入庫：担当者 
            destination TEXT,       -- 出庫：納品先
            requester TEXT,         -- 出庫：発注者
            admin_handler TEXT,     -- 出庫：総務課納品担当者
            memo TEXT,
            FOREIGN KEY(item_id) REFERENCES items(id)
        );
        """)
        self.conn.commit()

    def close(self):
        self.conn.close()

    def get_next_code(self) -> str:
        """
        備品コード：
        - 数字のみ
        - 10001 から開始
        - 欠番があれば最小の欠番を採番
        """
        cur = self.conn.cursor()
        cur.execute("""
            SELECT code FROM items 
            WHERE code GLOB '[0-9]*';
        """)
        used = set()
        for row in cur.fetchall():
            try:
                used.add(int(row["code"]))
            except:
                pass

        n = 10001
        while n in used:
            n += 1
        return str(n)

    def deactivate_item_free_code(self, item_id: int):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT code FROM items 
            WHERE id=?;
        """, (item_id,))
        row = cur.fetchone()
        if not row:
            raise ValueError("item not found")

        old_code = row["code"]
        date_str = datetime.now().strftime("%Y%m")
        retired_code = f"X{old_code}_{date_str}廃止"

        cur.execute("""
            UPDATE items
            SET is_active=0,
                code=?
            WHERE id=?;
        """, (retired_code, item_id))
        self.conn.commit()

    def add_item(self, code: str, name: str, location: str, unit: str, safety_stock: int, note: str):
        cur = self.conn.cursor()
        cur.execute("""
            INSERT INTO items (code, name, location, unit, safety_stock, note, is_active)
            VALUES (?, ?, ?, ?, ?, ?, 1);
        """, (code, name, location, unit, safety_stock, note))
        self.conn.commit()

    def update_item(self, item_id: int, code: str, name: str, location: str, unit: str,
                    safety_stock: int, note: str, is_active: int = 1):
        cur = self.conn.cursor()
        cur.execute("""
            UPDATE items
            SET code=?,
                name=?,
                location=?,
                unit=?,
                safety_stock=?,
                note=?,
                is_active=?
            WHERE id=?;
        """, (code, name, location, unit, safety_stock, note, is_active, item_id))
        self.conn.commit()

    def get_item_by_code(self, code):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT i.*, COALESCE(s.qty,0) as qty
            FROM items i
            LEFT JOIN stock s ON s.item_id=i.id
            WHERE i.code=? AND i.is_active=1;
        """, (code,))
        return cur.fetchone()

    def list_items(self, keyword: str = ""):
        kw = f"%{(keyword or '').strip()}%"
        cur = self.conn.cursor()
        cur.execute("""
            SELECT i.*, COALESCE(s.qty, 0) AS qty
            FROM items i
            LEFT JOIN (
                SELECT item_id, SUM(CASE WHEN type='IN' THEN qty ELSE -qty END) AS qty
                FROM transactions
                GROUP BY item_id
            ) s ON s.item_id = i.id
            WHERE i.is_active=1
              AND (
                    i.code LIKE ?
                 OR i.name LIKE ?
                 OR COALESCE(i.location,'') LIKE ?
                 OR COALESCE(i.unit,'') LIKE ?
                 OR COALESCE(i.note,'') LIKE ?
              )
            ORDER BY CAST(i.code AS INTEGER) ASC;
        """, (kw, kw, kw, kw, kw))
        return cur.fetchall()

    def list_transactions_by_type(self, tx_type: str, keyword: str = "", limit: int = 5000,
                                  start_ts: str | None = None, end_ts: str | None = None):
        kw = f"%{(keyword or '').strip()}%"
        cur = self.conn.cursor()

        where_ts = ""
        params = [tx_type]

        if start_ts and end_ts:
            where_ts = " AND t.ts >= ? AND t.ts < ? "
            params += [start_ts, end_ts]

        params += [
            kw, kw, kw, kw,     # i.code, i.name, i.location, i.note
            kw, kw,             # t.user, t.destination
            kw, kw, kw, kw,     # t.memo, t.supplier, t.requester, t.admin_handler
            int(limit)
        ]

        cur.execute(f"""
            SELECT t.*, i.code, i.name, i.unit
            FROM transactions t
            JOIN items i ON i.id = t.item_id
            WHERE t.type = ?
              {where_ts}
              AND (
                    i.code LIKE ?
                 OR i.name LIKE ?
                 OR COALESCE(i.location,'') LIKE ?
                 OR COALESCE(i.note,'') LIKE ?
                 OR COALESCE(t.user,'') LIKE ?
                 OR COALESCE(t.destination,'') LIKE ?
                 OR COALESCE(t.memo,'') LIKE ?
                 OR COALESCE(t.supplier,'') LIKE ?
                 OR COALESCE(t.requester,'') LIKE ?
                 OR COALESCE(t.admin_handler,'') LIKE ?
              )
            ORDER BY t.ts DESC
            LIMIT ?;
        """, tuple(params))

        return cur.fetchall()

    def _update_stock(self, item_id: int, delta: int):
        cur = self.conn.cursor()
        cur.execute("INSERT OR IGNORE INTO stock (item_id, qty) VALUES (?, 0);", (item_id,))
        cur.execute("UPDATE stock SET qty = qty + ? WHERE item_id=?;", (delta, item_id))
        self.conn.commit()

    def add_in_tx(self, item_id: int, qty: int, supplier: str, user: str, memo: str):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        cur = self.conn.cursor()
        cur.execute("""
            INSERT INTO transactions
            (ts, type, item_id, qty, supplier, user, memo)
            VALUES (?, 'IN', ?, ?, ?, ?, ?);
        """, (ts, item_id, qty, supplier, user, memo))
        self.conn.commit()

    def add_out_tx(self, item_id: int, qty: int, destination: str, requester: str, admin_handler: str, memo: str):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        cur = self.conn.cursor()
        cur.execute("""
            INSERT INTO transactions
            (ts, type, item_id, qty, destination, requester, admin_handler, memo)
            VALUES (?, 'OUT', ?, ?, ?, ?, ?, ?);
        """, (ts, item_id, qty, destination, requester, admin_handler, memo))
        self.conn.commit()

    def in_stock(self, item_id: int, qty: int, supplier: str, user: str, memo: str):
        self._update_stock(item_id, qty)
        self.add_in_tx(item_id, qty, supplier, user, memo)

    def out_stock(self, item_id: int, qty: int, destination: str, requester: str, admin_handler: str, memo: str):
        self._update_stock(item_id, -qty)
        self.add_out_tx(item_id, qty, destination, requester, admin_handler, memo)


# ======== バーコード作り ========
def ensure_dirs():
    os.makedirs(LABEL_DIR, exist_ok=True)

def generate_barcode_png(code: str) -> str:
    ensure_dirs()
    cls = barcode.get_barcode_class("code128")
    bc = cls(code, writer=ImageWriter())
    out_base = os.path.join(LABEL_DIR, f"{code}_barcode")
    filename = bc.save(out_base)
    return filename

def compose_label_png(code: str, name: str) -> str:
    """
    Create a simple label image:
      - item name
      - barcode
    """
    ensure_dirs()
    barcode_png = generate_barcode_png(code)

    W, H = 800, 400
    img = Image.new("RGB", (W, H), "white")
    draw = ImageDraw.Draw(img)

    font_big = None
    for f in ["meiryo.ttc", "MSYH.TTC", "arial.ttf"]:
        try:
            font_big = ImageFont.truetype(f, 44)
            break
        except:
            continue
    if font_big is None:
        font_big = ImageFont.load_default()

    # 備品名
    draw.text((20, 15), name, fill="black", font=font_big)

    # バーコード
    bc_img = Image.open(barcode_png).convert("RGB")
    bc_img = bc_img.resize((760, 300))
    img.paste(bc_img, (20, 110))

    out_path = os.path.join(LABEL_DIR, f"{code}_label.png")
    img.save(out_path)
    return out_path


# ======== UIサポート ========
def qitem(text: str) -> QTableWidgetItem:
    it = QTableWidgetItem(text if text is not None else "")
    it.setFlags(it.flags() ^ Qt.ItemIsEditable)
    return it

def warn(parent, title, msg):
    QMessageBox.warning(parent, title, msg)

def info(parent, title, msg):
    QMessageBox.information(parent, title, msg)


# ======== 操作画面 ========
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.db = DB(DB_FILE)

        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        # タブ
        self.tab_stock = QWidget()
        self.tab_in = QWidget()
        self.tab_out = QWidget()
        self.tab_master = QWidget()
        self.tab_in_history = QWidget()
        self.tab_out_history = QWidget()

        self.tabs.addTab(self.tab_stock, "在庫一覧")
        self.tabs.addTab(self.tab_in, "入庫")
        self.tabs.addTab(self.tab_out, "出庫")
        self.tabs.addTab(self.tab_master, "備品マスタ")
        self.tabs.addTab(self.tab_in_history, "入庫履歴")
        self.tabs.addTab(self.tab_out_history, "出庫履歴")

        self._build_stock_tab()
        self._build_in_tab()
        self._build_out_tab()
        self._build_master_tab()
        self._build_in_history_tab()
        self._build_out_history_tab()

        self.refresh_all()

    def _get_period_range(self, year_text: str, month_text: str):
        if year_text == "全部":
            return "ALL", None, None

        year = int(year_text)

        if month_text == "全部":
            start = f"{year:04d}-01-01 00:00"
            end = f"{year + 1:04d}-01-01 00:00"
            return f"{year:04d}", start, end

        month = int(month_text)
        start = f"{year:04d}-{month:02d}-01 00:00"
        if month == 12:
            end = f"{year + 1:04d}-01-01 00:00"
        else:
            end = f"{year:04d}-{month + 1:02d}-01 00:00"
        return f"{year:04d}-{month:02d}", start, end

    def closeEvent(self, event):
        try:
            self.db.close()
        except:
            pass
        super().closeEvent(event)

    def refresh_all(self):
        self.refresh_stock_list()
        self.refresh_in_history()
        self.refresh_out_history()

    # 在庫一覧
    def _build_stock_tab(self):
        layout = QVBoxLayout()

        # 検索欄
        top = QHBoxLayout()
        self.stock_search = QLineEdit()
        self.stock_search.setPlaceholderText("コード / 備品名 / 保管場所で検索")
        btn_search = QPushButton("検索")
        btn_search.clicked.connect(self.refresh_stock_list)
        self.stock_search.returnPressed.connect(self.refresh_stock_list)

        # ボタン
        btn_csv = QPushButton("CSV出力")
        btn_xlsx = QPushButton("Excel出力")
        btn_csv.clicked.connect(self.export_stock_csv)
        btn_xlsx.clicked.connect(self.export_stock_excel)

        top.addWidget(self.stock_search)
        top.addWidget(btn_search)
        top.addStretch()
        top.addWidget(btn_csv)
        top.addWidget(btn_xlsx)

        # 表示順
        self.stock_table = QTableWidget(0, 6)
        self.stock_table.setHorizontalHeaderLabels(
            ["コード", "備品名", "保管場所", "在庫数", "安全在庫", "状態"]
        )
        self.stock_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.stock_table.setEditTriggers(QTableWidget.NoEditTriggers)

        layout.addLayout(top)
        layout.addWidget(self.stock_table)
        self.tab_stock.setLayout(layout)

    def refresh_stock_list(self):
        keyword = self.stock_search.text().strip()
        rows = self.db.list_items(keyword)

        self.stock_table.setRowCount(0)

        for r in rows:
            row = self.stock_table.rowCount()
            self.stock_table.insertRow(row)

            code = r["code"]
            name = r["name"]
            location = r["location"] or ""
            unit = r["unit"] or ""

            qty = int(r["qty"])
            safety = int(r["safety_stock"] or 0)

            qty_text = f"{qty} {unit}".strip()
            safety_text = f"{safety} {unit}".strip()

            status = "OK" if qty >= safety else "不足"

            self.stock_table.setItem(row, 0, qitem(code))
            self.stock_table.setItem(row, 1, qitem(name))
            self.stock_table.setItem(row, 2, qitem(location))
            self.stock_table.setItem(row, 3, qitem(qty_text))
            self.stock_table.setItem(row, 4, qitem(safety_text))

            status_item = qitem(status)
            if status == "不足":
                status_item.setForeground(Qt.red)
            self.stock_table.setItem(row, 5, status_item)

        self.stock_table.resizeColumnsToContents()

    def export_stock_csv(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "CSV出力", "stock.csv", "CSV Files (*.csv)"
        )
        if not path:
            return

        import csv

        keyword = self.stock_search.text().strip()
        rows = self.db.list_items(keyword)

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow([
                "コード", "備品名", "保管場所",
                "在庫数", "単位",
                "安全在庫", "状態"
            ])

            for r in rows:
                qty = int(r["qty"])
                safety = int(r["safety_stock"] or 0)
                unit = r["unit"] or ""
                status = "OK" if qty >= safety else "不足"

                w.writerow([
                    r["code"],
                    r["name"],
                    r["location"] or "",
                    qty,
                    unit,
                    safety,
                    status
                ])

        info(self, "完了", "CSVを出力しました。")

    def export_stock_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel出力", "stock.xlsx", "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "在庫一覧"

        keyword = self.stock_search.text().strip()
        rows = self.db.list_items(keyword)

        headers = [
            "コード", "備品名", "保管場所",
            "在庫数", "単位",
            "安全在庫", "状態"
        ]
        ws.append(headers)

        for r in rows:
            qty = int(r["qty"])
            safety = int(r["safety_stock"] or 0)
            unit = r["unit"] or ""
            status = "OK" if qty >= safety else "不足"

            ws.append([
                r["code"],
                r["name"],
                r["location"] or "",
                qty,
                unit,
                safety,
                status
            ])

        widths = [14, 26, 20, 10, 8, 10, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(path)
        info(self, "完了", "Excelを出力しました。")

    # 入庫
    def _build_in_tab(self):
        layout = QVBoxLayout()
        form = QFormLayout()

        # -- コード入力 --
        self.in_code = QLineEdit()
        self.in_code.returnPressed.connect(self._in_load_item)

        # -- 備品情報表示 --
        self.in_name = QLabel("-")
        self.in_stock = QLabel("-")

        self.in_qty = QSpinBox()
        self.in_qty.setRange(1, 100000)
        self.in_qty.setValue(1)
        self.in_unit_label = QLabel("")

        qty_row = QHBoxLayout()
        qty_row.setContentsMargins(0, 0, 0, 0)
        qty_row.addWidget(self.in_qty)
        qty_row.addWidget(self.in_unit_label)
        qty_row.addStretch()

        # -- 購入先 / 担当者 / メモ --
        self.in_supplier = QLineEdit()
        self.in_user = QLineEdit()
        self.in_memo = QTextEdit()
        self.in_memo.setPlaceholderText("任意")

        # -- ボタン --
        btn_row = QHBoxLayout()

        dummy = QLabel()
        dummy.setFixedWidth(155)
        btn_row.addWidget(dummy)

        btn_load = QPushButton("読込")
        btn_save = QPushButton("入庫登録")
        btn_load.clicked.connect(self._in_load_item)
        btn_save.clicked.connect(self._do_in)

        btn_row.addWidget(btn_load)
        btn_row.addWidget(btn_save)
        btn_row.addStretch()

        # -- 表示順 --
        form.addRow("コード（スキャン）", self.in_code)
        form.addRow("備品名", self.in_name)
        form.addRow("現在庫", self.in_stock)
        form.addRow("数量", qty_row)
        form.addRow("購入先", self.in_supplier)
        form.addRow("担当者", self.in_user)
        form.addRow("メモ", self.in_memo)

        layout.addLayout(form)
        layout.addLayout(btn_row)
        layout.addStretch()
        self.tab_in.setLayout(layout)

    def _in_load_item(self):
        code = self.in_code.text().strip()
        if not code:
            return
        item = self.db.get_item_by_code(code)
        if not item:
            warn(self, "未登録", "このコードは備品マスタに存在しません。")
            self.in_name.setText("-")
            self.in_stock.setText("-")
            return
        self.in_name.setText(item["name"])
        unit = item["unit"] or ""
        self.in_stock.setText(f'{int(item["qty"])} {unit}'.strip())
        self._in_item_id = int(item["id"])
        self.in_unit_label.setText(item["unit"] or "")

    def _do_in(self):
        code = self.in_code.text().strip()
        if not code:
            warn(self, "入力エラー", "コードを入力してください。")
            return
        item = self.db.get_item_by_code(code)
        if not item:
            warn(self, "未登録", "このコードは備品マスタに存在しません。")
            return
        qty = int(self.in_qty.value())
        supplier = self.in_supplier.text().strip()
        user = self.in_user.text().strip()
        memo = self.in_memo.toPlainText().strip()

        self.db.in_stock(int(item["id"]), qty, supplier, user, memo)
        info(self, "完了", "入庫登録しました。")

        # -- 入力内容クリア --
        self.in_code.clear()
        self.in_supplier.clear()
        self.in_user.clear()
        self.in_memo.clear()
        self.in_qty.setValue(1)
        self.in_name.setText("-")
        self.in_stock.setText("-")
        self.refresh_all()
        self.in_code.setFocus()

    # 出庫
    def _build_out_tab(self):
        layout = QVBoxLayout()
        form = QFormLayout()

        # -- コード入力 --
        self.out_code = QLineEdit()
        self.out_code.returnPressed.connect(self._out_load_item)

        # -- 備品情報表示 --
        self.out_name = QLabel("-")
        self.out_stock = QLabel("-")
        self.out_safety = QLabel("-")

        self.out_qty = QSpinBox()
        self.out_qty.setRange(1, 100000)
        self.out_qty.setValue(1)

        self.out_unit_label = QLabel("")
        qty_row = QHBoxLayout()
        qty_row.setContentsMargins(0, 0, 0, 0)
        qty_row.addWidget(self.out_qty)
        qty_row.addWidget(self.out_unit_label)
        qty_row.addStretch()

        # -- 納品先 / 発注者 / 総務課納品担当者 / メモ --
        self.out_destination = QLineEdit()
        self.out_requester = QLineEdit()
        self.out_admin_handler = QLineEdit()
        self.out_memo = QTextEdit()
        self.out_memo.setPlaceholderText("任意")

        # -- ボタン --
        btn_row = QHBoxLayout()

        dummy = QLabel()
        dummy.setFixedWidth(155)
        btn_row.addWidget(dummy)

        btn_load = QPushButton("読込")
        btn_save = QPushButton("出庫登録")
        btn_load.clicked.connect(self._out_load_item)
        btn_save.clicked.connect(self._do_out)

        btn_row.addWidget(btn_load)
        btn_row.addWidget(btn_save)
        btn_row.addStretch()

        # -- 表示順 --
        form.addRow("コード（スキャン）", self.out_code)
        form.addRow("備品名", self.out_name)
        form.addRow("現在庫", self.out_stock)
        form.addRow("安全在庫", self.out_safety)
        form.addRow("数量", qty_row)
        form.addRow("納品先", self.out_destination)
        form.addRow("発注者", self.out_requester)
        form.addRow("総務課納品担当者", self.out_admin_handler)
        form.addRow("メモ", self.out_memo)

        layout.addLayout(form)
        layout.addLayout(btn_row)
        layout.addStretch()
        self.tab_out.setLayout(layout)

    def _out_load_item(self):
        code = self.out_code.text().strip()
        if not code:
            return
        item = self.db.get_item_by_code(code)
        if not item:
            warn(self, "未登録", "このコードは備品マスタに存在しません。")
            self.out_name.setText("-")
            self.out_stock.setText("-")
            self.out_safety.setText("-")
            return
        self.out_name.setText(item["name"])
        unit = item["unit"] or ""
        self.out_stock.setText(f'{int(item["qty"])} {unit}'.strip())
        unit = item["unit"] or ""
        safety = int(item["safety_stock"] or 0)
        self.out_safety.setText(f"{safety} {unit}".strip())
        self._out_item_id = int(item["id"])
        self.out_unit_label.setText(item["unit"] or "")

    def _do_out(self):
        code = self.out_code.text().strip()
        if not code:
            warn(self, "入力エラー", "コードを入力してください。")
            return
        item = self.db.get_item_by_code(code)
        if not item:
            warn(self, "未登録", "このコードは備品マスタに存在しません。")
            return

        qty = int(self.out_qty.value())
        now_qty = int(item["qty"]) if "qty" in item.keys() else 0
        safety = int(item["safety_stock"] or 0)
        if now_qty - qty < 0:
            ret = QMessageBox.question(
                self,
                "在庫不足",
                f"現在庫={now_qty}、出庫={qty} → 在庫がマイナスになります。\n強制出庫しますか？",
                QMessageBox.Yes | QMessageBox.No
            )
            if ret != QMessageBox.Yes:
                return
        elif now_qty - qty < safety:
            QMessageBox.information(
                self,
                "注意",
                f"出庫後の在庫が安全在庫を下回ります。（安全在庫={safety}）"
            )

        destination = self.out_destination.text().strip()
        requester = self.out_requester.text().strip()
        admin_handler = self.out_admin_handler.text().strip()
        memo = self.out_memo.toPlainText().strip()
        self.db.out_stock(int(item["id"]), qty, destination, requester, admin_handler, memo)
        info(self, "完了", "出庫登録しました。")

        # -- 入力内容クリア --
        self.out_code.clear()
        self.out_destination.clear()
        self.out_requester.clear()
        self.out_admin_handler.clear()
        self.out_memo.clear()
        self.out_qty.setValue(1)
        self.out_name.setText("-")
        self.out_stock.setText("-")
        self.out_safety.setText("-")
        self.out_unit_label.setText("")
        self.refresh_all()
        self.out_code.setFocus()

    # 備品マスタ
    def _build_master_tab(self):
        layout = QVBoxLayout()
        form = QFormLayout()

        # -- コード入力 --
        self.master_code = QLineEdit()
        self.master_code.setPlaceholderText("編集はコードで検索")
        self.master_code.returnPressed.connect(self.master_find)

        # -- 検索 / 新規ボタン --
        code_row = QHBoxLayout()
        btn_find = QPushButton("検索")
        btn_new = QPushButton("新規（自動採番）")
        btn_find.clicked.connect(self.master_find)
        btn_new.clicked.connect(self.master_new)

        code_row.setContentsMargins(0, 0, 0, 0)
        code_row.addWidget(self.master_code)
        code_row.addWidget(btn_find)
        code_row.addWidget(btn_new)
        code_row.addStretch()
        form.addRow("コード", code_row)

        # -- 備品情報表示 --
        self.master_name = QLineEdit()
        self.master_location = QLineEdit()
        self.master_unit = QLineEdit()
        self.master_safety = QSpinBox()
        self.master_safety.setRange(0, 100000)
        self.master_safety.setFixedWidth(120)
        self.master_note = QTextEdit()

        # -- 表示順 --
        form.addRow("備品名", self.master_name)
        form.addRow("保管場所", self.master_location)
        form.addRow("単位", self.master_unit)
        form.addRow("安全在庫", self.master_safety)
        form.addRow("メモ", self.master_note)

        # -- ボタン --
        btn_row = QHBoxLayout()

        dummy = QLabel()
        dummy.setFixedWidth(70)
        btn_row.addWidget(dummy)

        btn_save = QPushButton("保存")
        btn_label = QPushButton("ラベル作成（PNG）")
        btn_del = QPushButton("削除")
        btn_save.clicked.connect(self.master_save)
        btn_label.clicked.connect(self.master_make_label)
        btn_del.clicked.connect(self.master_delete)

        btn_row.addWidget(btn_save)
        btn_row.addWidget(btn_label)
        btn_row.addWidget(btn_del)
        btn_row.addStretch()

        layout.addLayout(form)
        layout.addLayout(btn_row)
        layout.addStretch()
        self.tab_master.setLayout(layout)

        self._master_item_id = None

    def master_new(self):
        try:
            code = self.db.get_next_code()
        except Exception as e:
            warn(self, "エラー", str(e))
            return
        self._master_item_id = None
        self.master_code.setText(code)
        self.master_name.clear()
        self.master_location.clear()
        self.master_unit.clear()
        self.master_safety.setValue(0)
        self.master_note.clear()
        self.master_name.setFocus()

    def master_find(self):
        code = self.master_code.text().strip()
        if not code:
            warn(self, "入力エラー", "検索するコードを入力してください。")
            return
        item = self.db.get_item_by_code(code)
        if not item:
            warn(self, "未登録", "該当する備品が見つかりません。")
            return
        self._master_item_id = int(item["id"])
        self.master_name.setText(item["name"])
        self.master_location.setText(item["location"] or "")
        self.master_unit.setText(item["unit"] or "")
        self.master_safety.setValue(int(item["safety_stock"] or 0))
        self.master_note.setPlainText(item["note"] or "")

    def master_save(self):
        code = self.master_code.text().strip()
        name = self.master_name.text().strip()
        if not code:
            warn(self, "入力エラー", "コードが空です。新規の場合は「新規（自動採番）」を押してください。")
            return
        if not name:
            warn(self, "入力エラー", "備品名は必須です。")
            return
        location = self.master_location.text().strip()
        unit = self.master_unit.text().strip()
        safety = int(self.master_safety.value())
        note = self.master_note.toPlainText().strip()

        try:
            if self._master_item_id is None:
                self.db.add_item(code, name, location, unit, safety, note)
                info(self, "完了", "新規備品を登録しました。")
            else:
                self.db.update_item(self._master_item_id, code, name, location, unit, safety, note, 1)
                info(self, "完了", "備品情報を更新しました。")
        except sqlite3.IntegrityError:
            warn(self, "エラー", "同じコードが既に存在します。")
            return

        self.refresh_all()

    def master_make_label(self):
        code = self.master_code.text().strip()
        name = self.master_name.text().strip()
        if not code or not name:
            warn(self, "入力エラー", "コードと備品名を入力してください。")
            return
        try:
            png = compose_label_png(code, name)
            info(self, "完了", f"ラベルを作成しました。\nPNG: {png}")
        except Exception as e:
            warn(self, "エラー", f"ラベル作成に失敗しました：{e}")

    def master_delete(self):
        if self._master_item_id is None:
            warn(self, "削除", "先にコードで検索して、削除する備品を表示してください。")
            return

        code = self.master_code.text().strip()
        name = self.master_name.text().strip()

        ret = QMessageBox.question(
            self,
            "削除確認",
            f"以下の備品を削除します。\n\nコード: {code}\n備品名: {name}\n\nよろしいですか？",
            QMessageBox.Yes | QMessageBox.No
        )
        if ret != QMessageBox.Yes:
            return

        try:
            self.db.deactivate_item_free_code(self._master_item_id)
        except Exception as e:
            warn(self, "エラー", f"削除に失敗しました：{e}")
            return

        info(self, "完了", "削除しました。番号は再利用できます。")

        # 画面クリア
        self._master_item_id = None
        self.master_code.clear()
        self.master_name.clear()
        self.master_location.clear()
        self.master_unit.clear()
        self.master_safety.setValue(0)
        self.master_note.clear()
        self.refresh_all()
        self.master_code.setFocus()

    # 入庫履歴
    def _build_in_history_tab(self):
        layout = QVBoxLayout()
        top = QHBoxLayout()

        # -- 年/月選択 --
        self.in_hist_year = QComboBox()
        self.in_hist_month = QComboBox()

        self.in_hist_year.addItem("全部")
        for y in range(2026, 2077):
            self.in_hist_year.addItem(str(y))

        self.in_hist_month.addItem("全部")
        for m in range(1, 13):
            self.in_hist_month.addItem(f"{m:02d}")

        def _in_year_changed():
            if self.in_hist_year.currentText() == "全部":
                self.in_hist_month.setCurrentText("全部")
            self.refresh_in_history()

        self.in_hist_year.currentIndexChanged.connect(_in_year_changed)
        self.in_hist_month.currentIndexChanged.connect(self.refresh_in_history)

        # -- 検索欄 --
        self.in_hist_search = QLineEdit()
        self.in_hist_search.setPlaceholderText("コード / 備品名 / 担当者 / 購入先で検索")
        self.in_hist_search.returnPressed.connect(self.refresh_in_history)

        # -- ボタン --
        btn_search = QPushButton("検索")
        btn_csv = QPushButton("CSV出力")
        btn_xlsx = QPushButton("Excel出力")
        btn_search.clicked.connect(self.refresh_in_history)
        btn_csv.clicked.connect(self.export_in_history_csv)
        btn_xlsx.clicked.connect(self.export_in_history_excel)

        top.addWidget(QLabel("年"))
        top.addWidget(self.in_hist_year)
        top.addWidget(QLabel("月"))
        top.addWidget(self.in_hist_month)

        top.addWidget(self.in_hist_search)
        top.addWidget(btn_search)
        top.addStretch()
        top.addWidget(btn_csv)
        top.addWidget(btn_xlsx)

        # -- 表示欄 --
        self.in_hist_table = QTableWidget(0, 8)
        self.in_hist_table.setHorizontalHeaderLabels(
            ["日時", "コード", "備品名", "数量", "単位", "購入先", "担当者", "メモ"]
        )
        self.in_hist_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.in_hist_table.setEditTriggers(QTableWidget.NoEditTriggers)

        layout.addLayout(top)
        layout.addWidget(self.in_hist_table)
        self.tab_in_history.setLayout(layout)

    def refresh_in_history(self):
        kw = self.in_hist_search.text().strip()

        period, start_ts, end_ts = self._get_period_range(
            self.in_hist_year.currentText(),
            self.in_hist_month.currentText()
        )

        rows = self.db.list_transactions_by_type("IN", kw, limit=5000, start_ts=start_ts, end_ts=end_ts)

        self.in_hist_table.setRowCount(0)
        for r in rows:
            row = self.in_hist_table.rowCount()
            self.in_hist_table.insertRow(row)

            unit = r["unit"] or ""
            qty = int(r["qty"])

            self.in_hist_table.setItem(row, 0, qitem(r["ts"]))
            self.in_hist_table.setItem(row, 1, qitem(r["code"]))
            self.in_hist_table.setItem(row, 2, qitem(r["name"]))
            self.in_hist_table.setItem(row, 3, qitem(str(qty)))
            self.in_hist_table.setItem(row, 4, qitem(unit))
            self.in_hist_table.setItem(row, 5, qitem(r["supplier"] or ""))
            self.in_hist_table.setItem(row, 6, qitem(r["user"] or ""))
            self.in_hist_table.setItem(row, 7, qitem(r["memo"] or ""))

        self.in_hist_table.resizeColumnsToContents()

    def export_in_history_csv(self):
        period, start_ts, end_ts = self._get_period_range(
            self.in_hist_year.currentText(),
            self.in_hist_month.currentText()
        )
        default_name = f"in_history_{period}.csv" if period != "ALL" else "in_history_ALL.csv"

        path, _ = QFileDialog.getSaveFileName(
            self, "入庫履歴CSV出力", default_name, "CSV Files (*.csv)"
        )
        if not path:
            return

        import csv

        kw = self.in_hist_search.text().strip()
        rows = self.db.list_transactions_by_type("IN", kw, limit=5000, start_ts=start_ts, end_ts=end_ts)

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["日時", "コード", "備品名", "数量", "単位", "購入先", "担当者", "メモ"])

            for r in rows:
                w.writerow([
                    r["ts"],
                    r["code"],
                    r["name"],
                    int(r["qty"]),
                    r["unit"] or "",
                    r["supplier"] or "",
                    r["user"] or "",
                    r["memo"] or ""
                ])

        info(self, "完了", "入庫履歴をCSV出力しました。")

    def export_in_history_excel(self):
        period, start_ts, end_ts = self._get_period_range(
            self.in_hist_year.currentText(),
            self.in_hist_month.currentText()
        )
        default_name = f"in_history_{period}.xlsx" if period != "ALL" else "in_history_ALL.xlsx"

        path, _ = QFileDialog.getSaveFileName(
            self, "入庫履歴Excel出力", default_name, "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "入庫履歴"

        headers = ["日時", "コード", "備品名", "数量", "単位", "購入先", "担当者", "メモ"]
        ws.append(headers)

        kw = self.in_hist_search.text().strip()
        rows = self.db.list_transactions_by_type("IN", kw, limit=5000, start_ts=start_ts, end_ts=end_ts)

        for r in rows:
            ws.append([
                r["ts"],
                r["code"],
                r["name"],
                int(r["qty"]),
                r["unit"] or "",
                r["supplier"] or "",
                r["user"] or "",
                r["memo"] or ""
            ])

        # -- 列の幅 --
        widths = [26, 14, 20, 8, 8, 18, 14, 26]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(path)
        info(self, "完了", "入庫履歴をExcel出力しました。")

    # 出庫履歴
    def _build_out_history_tab(self):
        layout = QVBoxLayout()
        top = QHBoxLayout()

        # -- 年/月選択 --
        self.out_hist_year = QComboBox()
        self.out_hist_month = QComboBox()

        self.out_hist_year.addItem("全部")
        for y in range(2026, 2077):
            self.out_hist_year.addItem(str(y))

        self.out_hist_month.addItem("全部")
        for m in range(1, 13):
            self.out_hist_month.addItem(f"{m:02d}")

        def _out_year_changed():
            if self.out_hist_year.currentText() == "全部":
                self.out_hist_month.setCurrentText("全部")
            self.refresh_out_history()

        self.out_hist_year.currentIndexChanged.connect(_out_year_changed)
        self.out_hist_month.currentIndexChanged.connect(self.refresh_out_history)

        # -- 検索欄 --
        self.out_hist_search = QLineEdit()
        self.out_hist_search.setPlaceholderText("コード / 備品名 / 納品先 / 発注者で検索")
        self.out_hist_search.returnPressed.connect(self.refresh_out_history)

        # -- ボタン --
        btn_search = QPushButton("検索")
        btn_csv = QPushButton("CSV出力")
        btn_xlsx = QPushButton("Excel出力")
        btn_search.clicked.connect(self.refresh_out_history)
        btn_csv.clicked.connect(self.export_out_history_csv)
        btn_xlsx.clicked.connect(self.export_out_history_excel)

        top.addWidget(QLabel("年"))
        top.addWidget(self.out_hist_year)
        top.addWidget(QLabel("月"))
        top.addWidget(self.out_hist_month)

        top.addWidget(self.out_hist_search)
        top.addWidget(btn_search)
        top.addStretch()
        top.addWidget(btn_csv)
        top.addWidget(btn_xlsx)

        # -- 表示欄 --
        self.out_hist_table = QTableWidget(0, 9)
        self.out_hist_table.setHorizontalHeaderLabels(
            ["日時", "コード", "備品名", "数量", "単位",
             "納品先", "発注者", "総務課納品担当者", "メモ"]
        )
        self.out_hist_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.out_hist_table.setEditTriggers(QTableWidget.NoEditTriggers)

        layout.addLayout(top)
        layout.addWidget(self.out_hist_table)
        self.tab_out_history.setLayout(layout)

    def refresh_out_history(self):
        kw = self.out_hist_search.text().strip()
        period, start_ts, end_ts = self._get_period_range(
            self.out_hist_year.currentText(),
            self.out_hist_month.currentText()
        )

        rows = self.db.list_transactions_by_type("OUT", kw, limit=5000, start_ts=start_ts, end_ts=end_ts)

        self.out_hist_table.setRowCount(0)
        for r in rows:
            row = self.out_hist_table.rowCount()
            self.out_hist_table.insertRow(row)

            unit = r["unit"] or ""

            self.out_hist_table.setItem(row, 0, qitem(r["ts"]))
            self.out_hist_table.setItem(row, 1, qitem(r["code"]))
            self.out_hist_table.setItem(row, 2, qitem(r["name"]))
            self.out_hist_table.setItem(row, 3, qitem(str(int(r["qty"]))))
            self.out_hist_table.setItem(row, 4, qitem(unit))
            self.out_hist_table.setItem(row, 5, qitem(r["destination"] or ""))  # 納品先
            self.out_hist_table.setItem(row, 6, qitem(r["requester"] or ""))  # 発注者
            self.out_hist_table.setItem(row, 7, qitem(r["admin_handler"] or ""))  # 総務課納品担当者
            self.out_hist_table.setItem(row, 8, qitem(r["memo"] or ""))

        self.out_hist_table.resizeColumnsToContents()

    def export_out_history_csv(self):
        period, start_ts, end_ts = self._get_period_range(
            self.out_hist_year.currentText(),
            self.out_hist_month.currentText()
        )
        default_name = f"out_history_{period}.csv" if period != "ALL" else "out_history_ALL.csv"

        path, _ = QFileDialog.getSaveFileName(
            self, "出庫履歴CSV出力", default_name, "CSV Files (*.csv)"
        )
        if not path:
            return

        import csv

        kw = self.out_hist_search.text().strip()
        rows = self.db.list_transactions_by_type("OUT", kw, limit=5000, start_ts=start_ts, end_ts=end_ts)

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["日時", "コード", "備品名", "数量", "単位", "納品先", "発注者", "総務課納品担当者", "メモ"])

            for r in rows:
                w.writerow([
                    r["ts"],
                    r["code"],
                    r["name"],
                    int(r["qty"]),
                    r["unit"] or "",
                    r["destination"] or "",  # 納品先
                    r["requester"] or "",  # 発注者
                    r["admin_handler"] or "",  # 総務課納品担当者
                    r["memo"] or ""
                ])

        info(self, "完了", "出庫履歴をCSV出力しました。")

    def export_out_history_excel(self):
        period, start_ts, end_ts = self._get_period_range(
            self.out_hist_year.currentText(),
            self.out_hist_month.currentText()
        )
        default_name = f"out_history_{period}.xlsx" if period != "ALL" else "out_history_ALL.xlsx"

        path, _ = QFileDialog.getSaveFileName(
            self, "出庫履歴Excel出力", default_name, "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "出庫履歴"

        headers = ["日時", "コード", "備品名", "数量", "単位", "納品先", "発注者", "総務課納品担当者", "メモ"]
        ws.append(headers)

        kw = self.out_hist_search.text().strip()
        rows = self.db.list_transactions_by_type("OUT", kw, limit=5000, start_ts=start_ts, end_ts=end_ts)

        for r in rows:
            ws.append([
                r["ts"],
                r["code"],
                r["name"],
                int(r["qty"]),
                r["unit"] or "",
                r["destination"] or "",  # 納品先
                r["requester"] or "",  # 発注者
                r["admin_handler"] or "",  # 総務課納品担当者
                r["memo"] or ""
            ])

        # 列幅（見やすく）
        widths = [26, 14, 20, 8, 8, 14, 14, 18, 26]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(path)
        info(self, "完了", "出庫履歴をExcel出力しました。")


def main():
    os.makedirs(LABEL_DIR, exist_ok=True)
    app = QApplication([])

    font = QFont()
    font.setFamily("Meiryo")
    font.setPointSize(13)
    app.setFont(font)

    win = MainWindow()
    win.resize(1100, 700)
    win.show()
    app.exec()


if __name__ == "__main__":
    main()


