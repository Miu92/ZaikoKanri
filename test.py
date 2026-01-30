import os
import sqlite3
from datetime import datetime
from dataclasses import dataclass

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QTabWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QMessageBox, QFormLayout, QSpinBox, QTextEdit, QComboBox, QFileDialog
)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Barcode + label
import barcode
from barcode.writer import ImageWriter
from PIL import Image, ImageDraw, ImageFont
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4


APP_TITLE = "備品在庫管理システム"
BASE_DIR = os.path.join(os.path.expanduser("~"), "Documents", "ZaikoKanri")
os.makedirs(BASE_DIR, exist_ok=True)
DB_FILE = os.path.join(BASE_DIR, "inventory.db")
LABEL_DIR = os.path.join(BASE_DIR, "labels")



# ---------------------------
# DB Layer
# ---------------------------
class DB:
    def __init__(self, path: str = DB_FILE):
        self.path = path
        self.conn = sqlite3.connect(self.path)
        self.conn.row_factory = sqlite3.Row
        self._init_schema()

    def _init_schema(self):
        cur = self.conn.cursor()
        cur.execute("""
        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL UNIQUE,
            name TEXT NOT NULL,
            category TEXT,
            location TEXT,
            unit TEXT,
            safety_stock INTEGER DEFAULT 0,
            note TEXT,
            is_active INTEGER DEFAULT 1
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS stock (
            item_id INTEGER PRIMARY KEY,
            qty INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY(item_id) REFERENCES items(id)
        );
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT NOT NULL,
            type TEXT NOT NULL, -- IN / OUT / ADJ
            item_id INTEGER NOT NULL,
            qty INTEGER NOT NULL,
            reason TEXT,
            user TEXT,
            memo TEXT,
            FOREIGN KEY(item_id) REFERENCES items(id)
        );
        """)
        self.conn.commit()

    def close(self):
        self.conn.close()

    def get_next_code(self) -> str:
        """
        備品コード：
        - 数字のみ
        - 10001 から開始
        """
        cur = self.conn.cursor()
        cur.execute("""
            SELECT code FROM items
            WHERE code GLOB '[0-9]*'
            ORDER BY CAST(code AS INTEGER) DESC
            LIMIT 1;
        """)
        row = cur.fetchone()

        if not row:
            return "10001"

        try:
            num = int(row["code"])
        except:
            num = 10000

        num += 1
        return str(num)

    def add_item(self, code, name, category, location, unit, safety_stock, note):
        cur = self.conn.cursor()
        cur.execute("""
            INSERT INTO items (code, name, category, location, unit, safety_stock, note, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?, 1);
        """, (code, name, category, location, unit, safety_stock, note))
        item_id = cur.lastrowid
        cur.execute("INSERT OR IGNORE INTO stock (item_id, qty) VALUES (?, 0);", (item_id,))
        self.conn.commit()
        return item_id

    def update_item(self, item_id, code, name, category, location, unit, safety_stock, note, is_active=1):
        cur = self.conn.cursor()
        cur.execute("""
            UPDATE items
            SET code=?, name=?, category=?, location=?, unit=?, safety_stock=?, note=?, is_active=?
            WHERE id=?;
        """, (code, name, category, location, unit, safety_stock, note, is_active, item_id))
        self.conn.commit()

    def get_item_by_code(self, code):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT i.*, COALESCE(s.qty,0) as qty
            FROM items i
            LEFT JOIN stock s ON s.item_id=i.id
            WHERE i.code=? AND i.is_active=1;
        """, (code,))
        return cur.fetchone()

    def list_items(self, keyword: str = ""):
        cur = self.conn.cursor()
        kw = f"%{keyword.strip()}%"
        cur.execute("""
            SELECT i.*, COALESCE(s.qty,0) as qty
            FROM items i
            LEFT JOIN stock s ON s.item_id=i.id
            WHERE i.is_active=1 AND (i.code LIKE ? OR i.name LIKE ?)
            ORDER BY i.code ASC;
        """, (kw, kw))
        return cur.fetchall()

    def list_transactions(self, keyword: str = "", limit: int = 2000):
        cur = self.conn.cursor()
        kw = f"%{keyword.strip()}%"
        cur.execute("""
            SELECT t.*, i.code, i.name
            FROM transactions t
            JOIN items i ON i.id=t.item_id
            WHERE (i.code LIKE ? OR i.name LIKE ? OR COALESCE(t.user,'') LIKE ? OR COALESCE(t.reason,'') LIKE ?)
            ORDER BY t.ts DESC
            LIMIT ?;
        """, (kw, kw, kw, kw, limit))
        return cur.fetchall()

    def list_transactions_by_type(self, tx_type: str, keyword: str = "", limit: int = 2000):
        cur = self.conn.cursor()
        kw = f"%{keyword.strip()}%"
        cur.execute("""
            SELECT t.*, i.code, i.name
            FROM transactions t
            JOIN items i ON i.id=t.item_id
            WHERE t.type=? AND
                  (i.code LIKE ? OR i.name LIKE ? OR COALESCE(t.user,'') LIKE ? OR COALESCE(t.reason,'') LIKE ?)
            ORDER BY t.ts DESC
            LIMIT ?;
        """, (tx_type, kw, kw, kw, kw, limit))
        return cur.fetchall()

    def _update_stock(self, item_id: int, delta: int):
        cur = self.conn.cursor()
        cur.execute("INSERT OR IGNORE INTO stock (item_id, qty) VALUES (?, 0);", (item_id,))
        cur.execute("UPDATE stock SET qty = qty + ? WHERE item_id=?;", (delta, item_id))
        self.conn.commit()

    def add_tx(self, tx_type: str, item_id: int, qty: int, reason: str, user: str, memo: str):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cur = self.conn.cursor()
        cur.execute("""
            INSERT INTO transactions (ts, type, item_id, qty, reason, user, memo)
            VALUES (?, ?, ?, ?, ?, ?, ?);
        """, (ts, tx_type, item_id, qty, reason, user, memo))
        self.conn.commit()

    def in_stock(self, item_id: int, qty: int, reason: str, user: str, memo: str):
        self._update_stock(item_id, qty)
        self.add_tx("IN", item_id, qty, reason, user, memo)

    def out_stock(self, item_id: int, qty: int, reason: str, user: str, memo: str):
        # qty is positive input, delta is -qty
        self._update_stock(item_id, -qty)
        self.add_tx("OUT", item_id, qty, reason, user, memo)


# ---------------------------
# Label & Barcode
# ---------------------------
def ensure_dirs():
    os.makedirs(LABEL_DIR, exist_ok=True)

def generate_barcode_png(code: str) -> str:
    ensure_dirs()
    # Code128 supports alphanumerics
    cls = barcode.get_barcode_class("code128")
    bc = cls(code, writer=ImageWriter())
    out_base = os.path.join(LABEL_DIR, f"{code}_barcode")
    filename = bc.save(out_base)  # generates ...png
    return filename

def compose_label_png(code: str, name: str) -> str:
    """
    Create a simple label image:
      - item name (large)
      - barcode
      - code text
    """
    ensure_dirs()
    barcode_png = generate_barcode_png(code)

    # Canvas size (pixels). Adjust as you like.
    W, H = 800, 400
    img = Image.new("RGB", (W, H), "white")
    draw = ImageDraw.Draw(img)

    # Try to load a font; fallback if not found
    font_big = None
    font_small = None
    for f in ["meiryo.ttc", "MSYH.TTC", "arial.ttf"]:
        try:
            font_big = ImageFont.truetype(f, 44)
            font_small = ImageFont.truetype(f, 28)
            break
        except:
            continue
    if font_big is None:
        font_big = ImageFont.load_default()
        font_small = ImageFont.load_default()

    # Name
    draw.text((20, 15), name, fill="black", font=font_big)

    # Barcode
    bc_img = Image.open(barcode_png).convert("RGB")
    # Resize barcode to fit
    bc_img = bc_img.resize((760, 180))
    img.paste(bc_img, (20, 110))

    # Code text
    draw.text((20, 310), f"コード: {code}", fill="black", font=font_small)

    out_path = os.path.join(LABEL_DIR, f"{code}_label.png")
    img.save(out_path)
    return out_path

# ---------------------------
# UI Helpers
# ---------------------------
def qitem(text: str) -> QTableWidgetItem:
    it = QTableWidgetItem(text if text is not None else "")
    it.setFlags(it.flags() ^ Qt.ItemIsEditable)
    return it

def warn(parent, title, msg):
    QMessageBox.warning(parent, title, msg)

def info(parent, title, msg):
    QMessageBox.information(parent, title, msg)


# ---------------------------
# Main Window
# ---------------------------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.db = DB(DB_FILE)

        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        # Tabs
        self.tab_stock = QWidget()
        self.tab_in = QWidget()
        self.tab_out = QWidget()
        self.tab_master = QWidget()
        self.tab_in_history = QWidget()
        self.tab_out_history = QWidget()

        self.tabs.addTab(self.tab_stock, "在庫一覧")
        self.tabs.addTab(self.tab_in, "入庫")
        self.tabs.addTab(self.tab_out, "出庫")
        self.tabs.addTab(self.tab_master, "備品マスタ")
        self.tabs.addTab(self.tab_in_history, "入庫履歴")
        self.tabs.addTab(self.tab_out_history, "出庫履歴")

        self._build_stock_tab()
        self._build_in_tab()
        self._build_out_tab()
        self._build_master_tab()
        self._build_in_history_tab()
        self._build_out_history_tab()

        self.refresh_all()

    def closeEvent(self, event):
        try:
            self.db.close()
        except:
            pass
        super().closeEvent(event)

    def refresh_all(self):
        self.refresh_stock_list()
        self.refresh_in_history()
        self.refresh_out_history()

    # ---- Tab: Stock list
    def _build_stock_tab(self):
        layout = QVBoxLayout()

        top = QHBoxLayout()
        self.stock_search = QLineEdit()
        self.stock_search.setPlaceholderText("コード / 備品名で検索")
        btn_search = QPushButton("検索")
        btn_search.clicked.connect(self.refresh_stock_list)
        self.stock_search.returnPressed.connect(self.refresh_stock_list)

        btn_csv = QPushButton("CSV出力")
        btn_xlsx = QPushButton("Excel出力")
        btn_csv.clicked.connect(self.export_stock_csv)
        btn_xlsx.clicked.connect(self.export_stock_excel)

        top.addWidget(self.stock_search)
        top.addWidget(btn_search)
        top.addStretch()
        top.addWidget(btn_csv)
        top.addWidget(btn_xlsx)

        self.stock_table = QTableWidget(0, 7)
        self.stock_table.setHorizontalHeaderLabels(
            ["コード", "備品名", "カテゴリ", "保管場所", "在庫数", "安全在庫", "状態"]
        )
        self.stock_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.stock_table.setEditTriggers(QTableWidget.NoEditTriggers)

        layout.addLayout(top)
        layout.addWidget(self.stock_table)
        self.tab_stock.setLayout(layout)

    def refresh_stock_list(self):
        keyword = self.stock_search.text().strip()
        rows = self.db.list_items(keyword)

        self.stock_table.setRowCount(0)
        for r in rows:
            row = self.stock_table.rowCount()
            self.stock_table.insertRow(row)

            qty = int(r["qty"])
            safety = int(r["safety_stock"] or 0)
            status = "OK" if qty >= safety else "不足"

            self.stock_table.setItem(row, 0, qitem(r["code"]))
            self.stock_table.setItem(row, 1, qitem(r["name"]))
            self.stock_table.setItem(row, 2, qitem(r["category"] or ""))
            self.stock_table.setItem(row, 3, qitem(r["location"] or ""))
            self.stock_table.setItem(row, 4, qitem(str(qty)))
            self.stock_table.setItem(row, 5, qitem(str(safety)))
            st_item = qitem(status)
            if status == "不足":
                st_item.setForeground(Qt.red)
            self.stock_table.setItem(row, 6, st_item)

        self.stock_table.resizeColumnsToContents()

    def export_stock_csv(self):
        path, _ = QFileDialog.getSaveFileName(self, "CSV出力", "stock.csv", "CSV Files (*.csv)")
        if not path:
            return
        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["コード", "備品名", "カテゴリ", "保管場所", "在庫数", "安全在庫", "状態"])
            for i in range(self.stock_table.rowCount()):
                row = [self.stock_table.item(i, c).text() for c in range(7)]
                w.writerow(row)
        info(self, "完了", "CSVを出力しました。")

    def export_stock_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Excel出力", "stock.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "在庫一覧"

        headers = ["コード", "備品名", "カテゴリ", "保管場所", "在庫数", "安全在庫", "状態"]
        ws.append(headers)
        for i in range(self.stock_table.rowCount()):
            ws.append([self.stock_table.item(i, c).text() for c in range(7)])

        # autosize
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18

        wb.save(path)
        info(self, "完了", "Excelを出力しました。")

    # ---- Tab: IN
    def _build_in_tab(self):
        layout = QVBoxLayout()
        form = QFormLayout()

        self.in_code = QLineEdit()
        self.in_code.setPlaceholderText("コードをスキャン（例：10001）")
        self.in_code.returnPressed.connect(self._in_load_item)

        self.in_name = QLabel("-")
        self.in_stock = QLabel("-")

        self.in_qty = QSpinBox()
        self.in_qty.setRange(1, 100000)
        self.in_qty.setValue(1)

        self.in_reason = QLineEdit()
        self.in_reason.setPlaceholderText("例：購⼊ / 返却")

        self.in_user = QLineEdit()
        self.in_user.setPlaceholderText("例：王 / 竹内")

        self.in_memo = QTextEdit()
        self.in_memo.setPlaceholderText("メモ（任意）")

        btn_row = QHBoxLayout()
        btn_load = QPushButton("読込")
        btn_save = QPushButton("入庫登録")
        btn_load.clicked.connect(self._in_load_item)
        btn_save.clicked.connect(self._do_in)

        btn_row.addWidget(btn_load)
        btn_row.addWidget(btn_save)
        btn_row.addStretch()

        form.addRow("コード（スキャン）", self.in_code)
        form.addRow("備品名", self.in_name)
        form.addRow("現在庫", self.in_stock)
        form.addRow("数量", self.in_qty)
        form.addRow("理由", self.in_reason)
        form.addRow("担当者", self.in_user)
        form.addRow("メモ", self.in_memo)

        layout.addLayout(form)
        layout.addLayout(btn_row)
        layout.addStretch()
        self.tab_in.setLayout(layout)

    def _in_load_item(self):
        code = self.in_code.text().strip()
        if not code:
            return
        item = self.db.get_item_by_code(code)
        if not item:
            warn(self, "未登録", "このコードは備品マスタに存在しません。")
            self.in_name.setText("-")
            self.in_stock.setText("-")
            return
        self.in_name.setText(item["name"])
        self.in_stock.setText(str(int(item["qty"])))
        self._in_item_id = int(item["id"])

    def _do_in(self):
        code = self.in_code.text().strip()
        if not code:
            warn(self, "入力エラー", "コードを入力（スキャン）してください。")
            return
        item = self.db.get_item_by_code(code)
        if not item:
            warn(self, "未登録", "このコードは備品マスタに存在しません。")
            return
        qty = int(self.in_qty.value())
        reason = self.in_reason.text().strip()
        user = self.in_user.text().strip()
        memo = self.in_memo.toPlainText().strip()

        self.db.in_stock(int(item["id"]), qty, reason, user, memo)
        info(self, "完了", "入庫登録しました。")
        self.in_code.clear()
        self.in_reason.clear()
        self.in_memo.clear()
        self.in_name.setText("-")
        self.in_stock.setText("-")
        self.refresh_all()
        self.in_code.setFocus()

    # ---- Tab: OUT
    def _build_out_tab(self):
        layout = QVBoxLayout()
        form = QFormLayout()

        self.out_code = QLineEdit()
        self.out_code.setPlaceholderText("コードをスキャン（例：10001）")
        self.out_code.returnPressed.connect(self._out_load_item)

        self.out_name = QLabel("-")
        self.out_stock = QLabel("-")
        self.out_safety = QLabel("-")

        self.out_qty = QSpinBox()
        self.out_qty.setRange(1, 100000)
        self.out_qty.setValue(1)

        self.out_reason = QLineEdit()
        self.out_reason.setPlaceholderText("例：納品 / 廃棄")

        self.out_user = QLineEdit()
        self.out_user.setPlaceholderText("例：総務課 / 和さび堂 / 大王庵")

        self.out_memo = QTextEdit()
        self.out_memo.setPlaceholderText("メモ（任意）")

        btn_row = QHBoxLayout()
        btn_load = QPushButton("読込")
        btn_save = QPushButton("出庫登録（強制可）")
        btn_load.clicked.connect(self._out_load_item)
        btn_save.clicked.connect(self._do_out)

        btn_row.addWidget(btn_load)
        btn_row.addWidget(btn_save)
        btn_row.addStretch()

        form.addRow("コード（スキャン）", self.out_code)
        form.addRow("備品名", self.out_name)
        form.addRow("現在庫", self.out_stock)
        form.addRow("安全在庫", self.out_safety)
        form.addRow("数量", self.out_qty)
        form.addRow("理由", self.out_reason)
        form.addRow("納品先", self.out_user)
        form.addRow("メモ", self.out_memo)

        layout.addLayout(form)
        layout.addLayout(btn_row)
        layout.addStretch()
        self.tab_out.setLayout(layout)

    def _out_load_item(self):
        code = self.out_code.text().strip()
        if not code:
            return
        item = self.db.get_item_by_code(code)
        if not item:
            warn(self, "未登録", "このコードは備品マスタに存在しません。")
            self.out_name.setText("-")
            self.out_stock.setText("-")
            self.out_safety.setText("-")
            return
        self.out_name.setText(item["name"])
        self.out_stock.setText(str(int(item["qty"])))
        self.out_safety.setText(str(int(item["safety_stock"] or 0)))
        self._out_item_id = int(item["id"])

    def _do_out(self):
        code = self.out_code.text().strip()
        if not code:
            warn(self, "入力エラー", "コードを入力（スキャン）してください。")
            return
        item = self.db.get_item_by_code(code)
        if not item:
            warn(self, "未登録", "このコードは備品マスタに存在しません。")
            return

        qty = int(self.out_qty.value())
        now_qty = int(item["qty"])
        safety = int(item["safety_stock"] or 0)

        # Force allowed: warn only
        if now_qty - qty < 0:
            ret = QMessageBox.question(
                self,
                "在庫不足（強制出庫）",
                f"現在庫={now_qty}、出庫={qty} → 在庫がマイナスになります。\n強制出庫しますか？",
                QMessageBox.Yes | QMessageBox.No
            )
            if ret != QMessageBox.Yes:
                return
        elif now_qty - qty < safety:
            # Not forbidden, just notice
            QMessageBox.information(
                self,
                "注意",
                f"出庫後の在庫が安全在庫を下回ります。（安全在庫={safety}）"
            )

        reason = self.out_reason.text().strip()
        user = self.out_user.text().strip()
        memo = self.out_memo.toPlainText().strip()

        self.db.out_stock(int(item["id"]), qty, reason, user, memo)
        info(self, "完了", "出庫登録しました。")
        self.out_code.clear()
        self.out_reason.clear()
        self.out_memo.clear()
        self.out_name.setText("-")
        self.out_stock.setText("-")
        self.out_safety.setText("-")
        self.refresh_all()
        self.out_code.setFocus()

    # ---- Tab: Master
    def _build_master_tab(self):
        layout = QVBoxLayout()

        top = QHBoxLayout()
        self.master_code = QLineEdit()
        self.master_code.setPlaceholderText("新規の場合は自動採番／編集はコードで検索")
        btn_find = QPushButton("検索")
        btn_new = QPushButton("新規（自動採番）")
        btn_save = QPushButton("保存")
        btn_label = QPushButton("ラベル作成（PNG）")

        btn_find.clicked.connect(self.master_find)
        btn_new.clicked.connect(self.master_new)
        btn_save.clicked.connect(self.master_save)
        btn_label.clicked.connect(self.master_make_label)

        top.addWidget(QLabel("コード"))
        top.addWidget(self.master_code)
        top.addWidget(btn_find)
        top.addStretch()
        top.addWidget(btn_new)
        top.addWidget(btn_save)
        top.addWidget(btn_label)

        form = QFormLayout()
        self.master_name = QLineEdit()
        self.master_category = QLineEdit()
        self.master_location = QLineEdit()
        self.master_unit = QLineEdit()
        self.master_safety = QSpinBox()
        self.master_safety.setRange(0, 100000)
        self.master_note = QTextEdit()

        form.addRow("備品名*", self.master_name)
        form.addRow("カテゴリ", self.master_category)
        form.addRow("保管場所", self.master_location)
        form.addRow("単位", self.master_unit)
        form.addRow("安全在庫", self.master_safety)
        form.addRow("メモ", self.master_note)

        layout.addLayout(top)
        layout.addLayout(form)
        layout.addStretch()
        self.tab_master.setLayout(layout)

        self._master_item_id = None

    def master_new(self):
        try:
            code = self.db.get_next_code()
        except Exception as e:
            warn(self, "エラー", str(e))
            return
        self._master_item_id = None
        self.master_code.setText(code)
        self.master_name.clear()
        self.master_category.clear()
        self.master_location.clear()
        self.master_unit.clear()
        self.master_safety.setValue(0)
        self.master_note.clear()
        self.master_name.setFocus()

    def master_find(self):
        code = self.master_code.text().strip()
        if not code:
            warn(self, "入力エラー", "検索するコードを入力してください。")
            return
        item = self.db.get_item_by_code(code)
        if not item:
            warn(self, "未登録", "該当する備品が見つかりません。")
            return
        self._master_item_id = int(item["id"])
        self.master_name.setText(item["name"])
        self.master_category.setText(item["category"] or "")
        self.master_location.setText(item["location"] or "")
        self.master_unit.setText(item["unit"] or "")
        self.master_safety.setValue(int(item["safety_stock"] or 0))
        self.master_note.setPlainText(item["note"] or "")

    def master_save(self):
        code = self.master_code.text().strip()
        name = self.master_name.text().strip()
        if not code:
            warn(self, "入力エラー", "コードが空です。新規の場合は「新規（自動採番）」を押してください。")
            return
        if not name:
            warn(self, "入力エラー", "備品名は必須です。")
            return
        category = self.master_category.text().strip()
        location = self.master_location.text().strip()
        unit = self.master_unit.text().strip()
        safety = int(self.master_safety.value())
        note = self.master_note.toPlainText().strip()

        try:
            if self._master_item_id is None:
                self.db.add_item(code, name, category, location, unit, safety, note)
                info(self, "完了", "新規備品を登録しました。")
            else:
                self.db.update_item(self._master_item_id, code, name, category, location, unit, safety, note, 1)
                info(self, "完了", "備品情報を更新しました。")
        except sqlite3.IntegrityError:
            warn(self, "エラー", "同じコードが既に存在します。")
            return

        self.refresh_all()

    def master_make_label(self):
        code = self.master_code.text().strip()
        name = self.master_name.text().strip()
        if not code or not name:
            warn(self, "入力エラー", "コードと備品名を入力してください。")
            return
        try:
            png = compose_label_png(code, name)
            info(self, "完了", f"ラベルを作成しました。\nPNG: {png}")
        except Exception as e:
            warn(self, "エラー", f"ラベル作成に失敗しました：{e}")

    # ---- Tab: in History
    def _build_in_history_tab(self):
        layout = QVBoxLayout()
        top = QHBoxLayout()

        self.in_hist_search = QLineEdit()
        self.in_hist_search.setPlaceholderText("コード / 備品名 / 担当者 / 理由 で検索")
        self.in_hist_search.returnPressed.connect(self.refresh_in_history)

        btn_search = QPushButton("検索")
        btn_search.clicked.connect(self.refresh_in_history)

        btn_csv = QPushButton("CSV出力")
        btn_csv.clicked.connect(self.export_in_history_csv)

        btn_xlsx = QPushButton("Excel出力")
        btn_xlsx.clicked.connect(self.export_in_history_excel)

        top.addWidget(self.in_hist_search)
        top.addWidget(btn_search)
        top.addStretch()
        top.addWidget(btn_csv)
        top.addWidget(btn_xlsx)

        self.in_hist_table = QTableWidget(0, 6)
        self.in_hist_table.setHorizontalHeaderLabels(
            ["日時", "コード", "備品名", "数量", "理由", "担当者"]
        )
        self.in_hist_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.in_hist_table.setEditTriggers(QTableWidget.NoEditTriggers)

        layout.addLayout(top)
        layout.addWidget(self.in_hist_table)
        self.tab_in_history.setLayout(layout)

    def refresh_in_history(self):
        kw = self.in_hist_search.text().strip()
        rows = self.db.list_transactions_by_type("IN", kw, limit=5000)

        self.in_hist_table.setRowCount(0)
        for r in rows:
            row = self.in_hist_table.rowCount()
            self.in_hist_table.insertRow(row)

            self.in_hist_table.setItem(row, 0, qitem(r["ts"]))
            self.in_hist_table.setItem(row, 1, qitem(r["code"]))
            self.in_hist_table.setItem(row, 2, qitem(r["name"]))
            self.in_hist_table.setItem(row, 3, qitem(str(int(r["qty"]))))
            self.in_hist_table.setItem(row, 4, qitem(r["reason"] or ""))
            self.in_hist_table.setItem(row, 5, qitem(r["user"] or ""))

        self.in_hist_table.resizeColumnsToContents()

    def export_in_history_csv(self):
        path, _ = QFileDialog.getSaveFileName(self, "入庫履歴CSV出力", "in_history.csv", "CSV Files (*.csv)")
        if not path:
            return
        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["日時", "コード", "備品名", "数量", "理由", "担当者"])
            for i in range(self.in_hist_table.rowCount()):
                w.writerow([self.in_hist_table.item(i, c).text() for c in range(6)])
        info(self, "完了", "入庫履歴をCSV出力しました。")

    def export_in_history_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "入庫履歴Excel出力", "in_history.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "入庫履歴"

        headers = ["日時", "コード", "備品名", "数量", "理由", "担当者"]
        ws.append(headers)

        for i in range(self.in_hist_table.rowCount()):
            ws.append([self.in_hist_table.item(i, c).text() for c in range(6)])

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

        wb.save(path)
        info(self, "完了", "入庫履歴をExcel出力しました。")

    # ---- Tab: in History
    def _build_out_history_tab(self):
        layout = QVBoxLayout()
        top = QHBoxLayout()

        self.out_hist_search = QLineEdit()
        self.out_hist_search.setPlaceholderText("コード / 備品名 / 納品先 / 理由 で検索")
        self.out_hist_search.returnPressed.connect(self.refresh_out_history)

        btn_search = QPushButton("検索")
        btn_search.clicked.connect(self.refresh_out_history)

        btn_csv = QPushButton("CSV出力")
        btn_csv.clicked.connect(self.export_out_history_csv)

        btn_xlsx = QPushButton("Excel出力")
        btn_xlsx.clicked.connect(self.export_out_history_excel)

        top.addWidget(self.out_hist_search)
        top.addWidget(btn_search)
        top.addStretch()
        top.addWidget(btn_csv)
        top.addWidget(btn_xlsx)

        self.out_hist_table = QTableWidget(0, 6)
        self.out_hist_table.setHorizontalHeaderLabels(
            ["日時", "コード", "備品名", "数量", "理由", "納品先"]
        )
        self.out_hist_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.out_hist_table.setEditTriggers(QTableWidget.NoEditTriggers)

        layout.addLayout(top)
        layout.addWidget(self.out_hist_table)
        self.tab_out_history.setLayout(layout)

    def refresh_out_history(self):
        kw = self.out_hist_search.text().strip()
        rows = self.db.list_transactions_by_type("OUT", kw, limit=5000)

        self.out_hist_table.setRowCount(0)
        for r in rows:
            row = self.out_hist_table.rowCount()
            self.out_hist_table.insertRow(row)

            self.out_hist_table.setItem(row, 0, qitem(r["ts"]))
            self.out_hist_table.setItem(row, 1, qitem(r["code"]))
            self.out_hist_table.setItem(row, 2, qitem(r["name"]))
            self.out_hist_table.setItem(row, 3, qitem(str(int(r["qty"]))))
            self.out_hist_table.setItem(row, 4, qitem(r["reason"] or ""))
            self.out_hist_table.setItem(row, 5, qitem(r["user"] or ""))

        self.out_hist_table.resizeColumnsToContents()

    def export_out_history_csv(self):
        path, _ = QFileDialog.getSaveFileName(self, "出庫履歴CSV出力", "out_history.csv", "CSV Files (*.csv)")
        if not path:
            return
        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["日時", "コード", "備品名", "数量", "理由", "納品先"])
            for i in range(self.out_hist_table.rowCount()):
                w.writerow([self.out_hist_table.item(i, c).text() for c in range(6)])
        info(self, "完了", "出庫履歴をCSV出力しました。")

    def export_out_history_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "出庫履歴Excel出力", "out_history.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "出庫履歴"

        headers = ["日時", "コード", "備品名", "数量", "理由", "納品先"]
        ws.append(headers)

        for i in range(self.out_hist_table.rowCount()):
            ws.append([self.out_hist_table.item(i, c).text() for c in range(6)])

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

        wb.save(path)
        info(self, "完了", "出庫履歴をExcel出力しました。")


def main():
    os.makedirs(LABEL_DIR, exist_ok=True)
    app = QApplication([])
    win = MainWindow()
    win.resize(1100, 700)
    win.show()
    app.exec()


if __name__ == "__main__":
    main()
